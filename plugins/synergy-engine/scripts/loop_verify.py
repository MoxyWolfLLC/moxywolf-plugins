"""
loop_verify.py — the seven deterministic checks the Synergy Engine prep
loop must pass after every tick. No LLM call. No network call except the
HEAD-200 citation check.

Run directly:

    python3 loop_verify.py /path/to/.loop-state /path/to/tracker.xlsx [--rows-written A,B,C]

Exit code 0 = all checks passed. Exit code 20 = hard stop (writes halt
reason into state). Exit code 30 = transient error. See
references/loop-contract.md for the full contract.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

# Stdlib HTTP, so the verifier has no third-party dep beyond openpyxl.
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen

try:
    import openpyxl  # type: ignore
except ImportError:
    print(
        "ERROR: openpyxl is required (pip install openpyxl)", file=sys.stderr
    )
    sys.exit(30)

import loop_state as ls

# --- Legal enum values, mirrored from references/tracker-schema.md ----------

EXPECTED_COLUMNS = [
    "Target",
    "Persona",
    "Tier",
    "Path",
    "Synergy",
    "LinkedIn Profile",
    "Last Touch",
    "Liked",
    "Commented",
    "Comment / engagement summary",
    "Cited /answers URL",
    "Connect / DM",
    "Status",
    "Next Action",
    "Next Action Date",
    "Notes",
]

# Loop-writable status values. Anything else in column M flips check #4.
LOOP_WRITABLE_STATUS = {"Not started", "Queued", "Ready for review"}

LEGAL_STATUS = LOOP_WRITABLE_STATUS | {
    "Posted",
    "Pending accept",
    "Accepted",
    "Replied",
    "Engaged",
    "Parked",
    "Cold",
}

LEGAL_TIER = {"A", "B", "C", "peer"}
LEGAL_PATH = {"A", "B", "GA", "cite-only", "cited"}
LEGAL_SYNERGY = {"High", "Medium", "Low", "Cold"}

# Persona is open-ended in the schema ("Content / Content/peer / Citation /
# GRC / Builder / Bridge"). We require non-empty but don't lock the enum.
PERSONA_REQUIRED = True

CITED_PATHS = {"A", "GA", "cited"}

# --- Result types ----------------------------------------------------------


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str = ""
    # If True, failure of this check is a hard stop, not a soft fail.
    hard_stop: bool = False


@dataclass
class VerifierReport:
    checks: list[CheckResult] = field(default_factory=list)
    made_progress: bool = False
    # §4 Governed Autonomy: the report carries the owner's name into every
    # audit record. The polish bias (§2) can't hide a missing name.
    release_owner: str = ""
    # Chain integrity. plan_nonce_echoed is the nonce the tick handed in;
    # the report has to echo it back verbatim or the tick refuses to record
    # this report's result. verifier_nonce is the nonce issued to THIS
    # verifier run by state.issue_nonce(NONCE_VERIFIER); the tick consumes it
    # after reading the report. Either side faking the other breaks the chain.
    plan_nonce_echoed: str = ""
    verifier_nonce: str = ""
    tick_id: str = ""
    utc: str = ""

    @property
    def all_passed(self) -> bool:
        return all(c.passed for c in self.checks)

    @property
    def hard_stop_reason(self) -> Optional[str]:
        for c in self.checks:
            if not c.passed and c.hard_stop:
                return f"{c.name}: {c.detail}"
        return None

    def to_dict(self) -> dict[str, Any]:
        return {
            "all_passed": self.all_passed,
            "made_progress": self.made_progress,
            "hard_stop_reason": self.hard_stop_reason,
            "release_owner": self.release_owner,
            "plan_nonce_echoed": self.plan_nonce_echoed,
            "verifier_nonce": self.verifier_nonce,
            "tick_id": self.tick_id,
            "utc": self.utc,
            "checks": [
                {
                    "name": c.name,
                    "passed": c.passed,
                    "detail": c.detail,
                    "hard_stop": c.hard_stop,
                }
                for c in self.checks
            ],
        }


# --- Helpers ---------------------------------------------------------------


def _load_tracker_rows(tracker_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    """Return (headers, rows-as-dicts) from the Outreach Tracker sheet."""
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True, read_only=True)
    if "Outreach Tracker" not in wb.sheetnames:
        raise ValueError("sheet 'Outreach Tracker' not found")
    ws = wb["Outreach Tracker"]
    iter_rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(iter_rows)]
    except StopIteration:
        return [], []
    rows = []
    for r in iter_rows:
        if all(v is None or v == "" for v in r):
            continue
        rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return headers, rows


def _head_ok(url: str, timeout: float = 5.0) -> bool:
    """HEAD-200 (or 2xx/3xx) check with one retry. Used by check #3."""
    if not isinstance(url, str) or not url.strip():
        return False
    if not re.match(r"^https?://", url.strip()):
        return False
    for attempt in (1, 2):
        try:
            req = Request(url, method="HEAD", headers={"User-Agent": "synergy-loop-verify/1"})
            with urlopen(req, timeout=timeout) as resp:
                if 200 <= resp.status < 400:
                    return True
        except (HTTPError, URLError, TimeoutError):
            if attempt == 2:
                return False
    return False


def _is_blank(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


# --- The seven checks ------------------------------------------------------


def check_1_tracker_integrity(tracker_path: Path) -> CheckResult:
    """#1: xlsx opens, sheet exists, columns A-P present, no duplicate
    (Target, LinkedIn Profile) rows."""
    try:
        headers, rows = _load_tracker_rows(tracker_path)
    except (FileNotFoundError, ValueError) as e:
        return CheckResult("1_tracker_integrity", False, str(e), hard_stop=True)
    missing = [c for c in EXPECTED_COLUMNS if c not in headers]
    if missing:
        return CheckResult(
            "1_tracker_integrity",
            False,
            f"missing columns: {missing}",
            hard_stop=True,
        )
    seen: set[tuple[str, str]] = set()
    dupes: list[tuple[str, str]] = []
    for r in rows:
        key = (
            str(r.get("Target") or "").strip().lower(),
            str(r.get("LinkedIn Profile") or "").strip().lower(),
        )
        if key == ("", ""):
            continue
        if key in seen:
            dupes.append(key)
        seen.add(key)
    if dupes:
        return CheckResult(
            "1_tracker_integrity",
            False,
            f"duplicate (Target, LinkedIn Profile) rows: {dupes[:3]}",
        )
    return CheckResult("1_tracker_integrity", True, f"{len(rows)} rows")


def check_2_schema_valid_writes(
    rows: list[dict[str, Any]], rows_written_keys: set[str]
) -> CheckResult:
    """#2: each row written this tick has required cols + legal enum values."""
    if not rows_written_keys:
        return CheckResult("2_schema_valid_writes", True, "no rows written this tick")
    bad: list[str] = []
    for r in rows:
        key = str(r.get("Target") or "").strip()
        if key not in rows_written_keys:
            continue
        required = ["Target", "Persona", "Tier", "Path", "Synergy",
                    "LinkedIn Profile", "Comment / engagement summary",
                    "Status", "Next Action", "Next Action Date"]
        for col in required:
            if _is_blank(r.get(col)):
                bad.append(f"{key}: missing {col}")
        if r.get("Path") in CITED_PATHS and _is_blank(r.get("Cited /answers URL")):
            bad.append(f"{key}: Path={r.get('Path')} requires Cited /answers URL")
        if not _is_blank(r.get("Tier")) and r.get("Tier") not in LEGAL_TIER:
            bad.append(f"{key}: illegal Tier={r.get('Tier')!r}")
        if not _is_blank(r.get("Path")) and r.get("Path") not in LEGAL_PATH:
            bad.append(f"{key}: illegal Path={r.get('Path')!r}")
        if not _is_blank(r.get("Synergy")) and r.get("Synergy") not in LEGAL_SYNERGY:
            bad.append(f"{key}: illegal Synergy={r.get('Synergy')!r}")
        if not _is_blank(r.get("Status")) and r.get("Status") not in LEGAL_STATUS:
            bad.append(f"{key}: illegal Status={r.get('Status')!r}")
        if PERSONA_REQUIRED and _is_blank(r.get("Persona")):
            bad.append(f"{key}: missing Persona")
    if bad:
        return CheckResult("2_schema_valid_writes", False, "; ".join(bad[:5]))
    return CheckResult(
        "2_schema_valid_writes", True, f"{len(rows_written_keys)} rows ok"
    )


def check_3_citations_verified(
    rows: list[dict[str, Any]], rows_written_keys: set[str]
) -> CheckResult:
    """#3: for each cited path row written this tick, Cited URL is HEAD-200
    and Comment / engagement summary is non-empty."""
    if not rows_written_keys:
        return CheckResult("3_citations_verified", True, "no rows written this tick")
    bad: list[str] = []
    checked = 0
    for r in rows:
        key = str(r.get("Target") or "").strip()
        if key not in rows_written_keys:
            continue
        if r.get("Path") not in CITED_PATHS:
            continue
        checked += 1
        url = str(r.get("Cited /answers URL") or "").strip()
        if _is_blank(r.get("Comment / engagement summary")):
            bad.append(f"{key}: empty engagement summary")
        if not _head_ok(url):
            bad.append(f"{key}: Cited /answers URL not HEAD-200: {url}")
    if bad:
        return CheckResult("3_citations_verified", False, "; ".join(bad[:5]))
    return CheckResult("3_citations_verified", True, f"{checked} cited rows ok")


def check_4_human_gate_untouched(
    rows: list[dict[str, Any]], rows_written_keys: set[str]
) -> CheckResult:
    """#4: the loop must not touch send-side columns or write a Status past
    Ready for review. THIS IS A HARD STOP and the loop refuses to resume
    once it fires."""
    if not rows_written_keys:
        return CheckResult(
            "4_human_gate_untouched", True, "no rows written this tick"
        )
    violations: list[str] = []
    forbidden_cols = ["Liked", "Commented", "Connect / DM", "Last Touch"]
    for r in rows:
        key = str(r.get("Target") or "").strip()
        if key not in rows_written_keys:
            continue
        for col in forbidden_cols:
            if not _is_blank(r.get(col)):
                violations.append(f"{key}: loop wrote {col}={r.get(col)!r}")
        status = r.get("Status")
        if not _is_blank(status) and status not in LOOP_WRITABLE_STATUS:
            violations.append(
                f"{key}: loop wrote Status={status!r} (only {LOOP_WRITABLE_STATUS} allowed)"
            )
    if violations:
        return CheckResult(
            "4_human_gate_untouched",
            False,
            "; ".join(violations[:5]),
            hard_stop=True,
        )
    return CheckResult("4_human_gate_untouched", True, "send side untouched")


def check_5_quota_envelope(state: ls.State) -> CheckResult:
    """#5: daily API counters under per-day ceilings."""
    over: list[str] = []
    if state.quotas.apify_calls_today > state.bounds.max_apify_per_day:
        over.append(
            f"apify {state.quotas.apify_calls_today}/{state.bounds.max_apify_per_day}"
        )
    if state.quotas.apollo_calls_today > state.bounds.max_apollo_per_day:
        over.append(
            f"apollo {state.quotas.apollo_calls_today}/{state.bounds.max_apollo_per_day}"
        )
    if state.quotas.openalex_calls_today > state.bounds.max_openalex_per_day:
        over.append(
            f"openalex {state.quotas.openalex_calls_today}/{state.bounds.max_openalex_per_day}"
        )
    if over:
        return CheckResult("5_quota_envelope", False, "; ".join(over))
    return CheckResult(
        "5_quota_envelope",
        True,
        f"apify={state.quotas.apify_calls_today} apollo={state.quotas.apollo_calls_today} openalex={state.quotas.openalex_calls_today}",
    )


def check_6_calendar_block_present(state: ls.State, step_just_ran: Optional[str] = None) -> CheckResult:
    """#6: if a batch is ready (drafts_ready_for_review >= n_per_batch),
    a calendar block must be booked. This is a structural check that
    fires at the BATCH level, not the tick level: a tick whose job was
    not booking the calendar should not be penalized for the absence of
    a block. We treat the gap as a pending state, not a failure, unless
    the BOOK_CALENDAR step itself just ran and didn't deliver a block."""
    ready = state.handoff.drafts_ready_for_review
    target = state.bounds.n_per_batch
    if ready < target:
        return CheckResult(
            "6_calendar_block_present",
            True,
            f"batch not ready yet ({ready}/{target}); no block required",
        )
    if not state.handoff.last_calendar_block_id:
        # Only a real failure when the step that JUST ran was supposed to
        # book the block. Otherwise it's a known-pending state the
        # dispatcher will address on the next BOOK_CALENDAR step.
        if step_just_ran == "BOOK_CALENDAR":
            return CheckResult(
                "6_calendar_block_present",
                False,
                f"batch ready ({ready}/{target}) but BOOK_CALENDAR step did not produce a block",
            )
        return CheckResult(
            "6_calendar_block_present",
            True,
            f"batch ready ({ready}/{target}); calendar block pending (dispatcher will plan BOOK_CALENDAR next)",
        )
    return CheckResult(
        "6_calendar_block_present",
        True,
        f"block {state.handoff.last_calendar_block_id} at {state.handoff.last_calendar_block_utc}",
    )


def check_7_progress_signal(
    state_before: ls.State, state_after: ls.State, fingerprint_refreshed: bool
) -> CheckResult:
    """#7: at least one of: queued grew, registry-pending shrank, drafts-ready
    grew, fingerprint refreshed."""
    queued_grew = (
        state_after.discovery.queued_since_last_run
        > state_before.discovery.queued_since_last_run
    )
    pending_shrank = (
        state_after.citation.registry_pending_verify
        < state_before.citation.registry_pending_verify
    )
    drafts_grew = (
        state_after.handoff.drafts_ready_for_review
        > state_before.handoff.drafts_ready_for_review
    )
    progress = any([queued_grew, pending_shrank, drafts_grew, fingerprint_refreshed])
    cap = state_after.bounds.max_consecutive_no_progress
    consecutive = state_after.loop.consecutive_no_progress_ticks
    if not progress and consecutive >= cap:
        return CheckResult(
            "7_progress_signal",
            False,
            f"no progress for {consecutive} consecutive ticks (cap {cap})",
            hard_stop=True,
        )
    return CheckResult(
        "7_progress_signal",
        True,
        (
            f"progress=True (queued_grew={queued_grew} pending_shrank={pending_shrank} "
            f"drafts_grew={drafts_grew} fp_refreshed={fingerprint_refreshed})"
            if progress
            else f"no progress this tick (streak {consecutive}/{cap})"
        ),
    )


# --- Driver ----------------------------------------------------------------


def run_all_checks(
    loop_state_dir: Path,
    tracker_path: Path,
    rows_written_keys: Optional[set[str]] = None,
    state_before: Optional[ls.State] = None,
    fingerprint_refreshed: bool = False,
    step_just_ran: Optional[str] = None,
    plan_nonce: Optional[str] = None,
) -> VerifierReport:
    """Run all seven checks. Caller passes rows_written_keys (Targets the
    tick just wrote) and the pre-tick state snapshot for the progress check.

    plan_nonce: when present, the verifier issues a fresh verifier_nonce
    bound to the active tick (via ls.issue_nonce) and echoes plan_nonce
    verbatim into the report. The tick's resolve step refuses to record any
    report whose plan_nonce_echoed doesn't match the nonce it handed in.
    Run-direct-from-CLI verifications can omit plan_nonce; those produce a
    report with empty nonce fields, which the tick will reject. That's the
    whole point: a hand-run verifier can't masquerade as a chained one."""
    rows_written_keys = rows_written_keys or set()
    state_after = ls.load(loop_state_dir)
    state_before = state_before or state_after

    # Issue our verifier nonce if and only if the caller presented a valid
    # plan_nonce. ls.issue_nonce raises if the nonce doesn't match, which
    # bubbles up and aborts the verifier run — by design.
    verifier_nonce = ""
    if plan_nonce:
        verifier_nonce = ls.issue_nonce(loop_state_dir, ls.NONCE_VERIFIER, plan_nonce)
        # Re-load state because issue_nonce mutated chain.
        state_after = ls.load(loop_state_dir)

    headers, rows = [], []
    integrity = check_1_tracker_integrity(tracker_path)
    if integrity.passed:
        # Already opened it inside check_1; cheap to do again with full read.
        try:
            _, rows = _load_tracker_rows(tracker_path)
        except Exception as e:
            integrity = CheckResult(
                "1_tracker_integrity", False, str(e), hard_stop=True
            )

    report = VerifierReport(
        release_owner=state_after.release_owner,
        plan_nonce_echoed=plan_nonce or "",
        verifier_nonce=verifier_nonce,
        tick_id=state_after.chain.tick_id or "",
        utc=ls.utcnow_iso(),
    )
    report.checks.append(integrity)
    if integrity.passed:
        report.checks.append(check_4_human_gate_untouched(rows, rows_written_keys))
        report.checks.append(check_2_schema_valid_writes(rows, rows_written_keys))
        report.checks.append(check_3_citations_verified(rows, rows_written_keys))
    report.checks.append(check_5_quota_envelope(state_after))
    report.checks.append(check_6_calendar_block_present(state_after, step_just_ran=step_just_ran))
    report.checks.append(
        check_7_progress_signal(state_before, state_after, fingerprint_refreshed)
    )
    report.made_progress = any(
        c.name == "7_progress_signal" and c.passed and "no progress" not in c.detail
        for c in report.checks
    )

    # Persist last verifier output for /synergy-loop-status.
    (loop_state_dir / "last_verifier.json").write_text(
        json.dumps(report.to_dict(), indent=2, sort_keys=True),
        encoding="utf-8",
    )

    # If check #4 fired, write the halt reason into state so loop_tick.py
    # refuses to resume.
    for c in report.checks:
        if not c.passed and c.hard_stop and c.name == "4_human_gate_untouched":
            state_after.loop.halt_reason = f"LOOP_VIOLATED_HUMAN_GATE: {c.detail}"
            ls.save(loop_state_dir, state_after)
            break

    return report


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("loop_state_dir", type=Path)
    p.add_argument("tracker_path", type=Path)
    p.add_argument(
        "--rows-written",
        default="",
        help="Comma-separated Target names written this tick. Empty for a dry verify.",
    )
    p.add_argument(
        "--fingerprint-refreshed",
        action="store_true",
        help="Pass when the tick refreshed the fingerprint.",
    )
    p.add_argument(
        "--plan-nonce",
        default="",
        help="The active tick's plan_nonce. When present, the verifier issues "
             "a verifier_nonce against the chain and echoes the plan_nonce "
             "into the report. Omit for a hand-run dry verification (whose "
             "report the tick will then refuse to record).",
    )
    p.add_argument(
        "--step-just-ran",
        default="",
        help="The step name the tick just resolved (e.g. BOOK_CALENDAR). "
             "Used by check #6 to scope the calendar-block requirement.",
    )
    args = p.parse_args(argv)

    rows_written = {
        s.strip() for s in args.rows_written.split(",") if s.strip()
    }

    try:
        report = run_all_checks(
            loop_state_dir=args.loop_state_dir,
            tracker_path=args.tracker_path,
            rows_written_keys=rows_written,
            fingerprint_refreshed=args.fingerprint_refreshed,
            step_just_ran=args.step_just_ran or None,
            plan_nonce=args.plan_nonce or None,
        )
    except ls.StateNotFoundError as e:
        print(f"ERROR: no state.json at {e}; run /synergy-loop-start first", file=sys.stderr)
        return 30
    except ls.StateSchemaError as e:
        print(f"ERROR: state.json schema invalid: {e}", file=sys.stderr)
        return 30

    print(json.dumps(report.to_dict(), indent=2, sort_keys=True))

    if report.hard_stop_reason:
        print(f"\nHARD STOP: {report.hard_stop_reason}", file=sys.stderr)
        return ls.EXIT_HARD_STOP
    if not report.all_passed:
        print("\nVERIFIER FAIL (soft); tick should be retried or rolled back", file=sys.stderr)
        return ls.EXIT_TRANSIENT
    return ls.EXIT_CONTINUE


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
