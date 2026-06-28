"""
loop_tick.py — one-unit-of-work dispatcher for the Synergy Engine prep loop.

Each cron firing is one tick. A tick:

  1. Loads state + checks halt + checks bounds
  2. Picks exactly ONE of the cursor steps below:
       - REFRESH_FINGERPRINT  (sources changed)
       - DISCOVER_AUTHOR      (author-center queue is thin)
       - DISCOVER_CONTENT     (content-center queue is thin)
       - HARVEST_CITATIONS    (bibliography changed or registry pending)
       - STAGE_DRAFTS         (we have enough queued; draft N rows)
       - BOOK_CALENDAR        (drafts ready; no calendar block booked yet)
       - IDLE                 (goal met or nothing actionable this tick)
  3. Emits a STEP PLAN as JSON on stdout, the agent reads it and runs the
     matching slash command. The agent then calls
       loop_tick.py --resolve <step> --result <ok|fail|...> --metric K=V ...
     to record what happened.
  4. Runs the verifier; updates state atomically; exits with the right code.

Two modes:

  python3 loop_tick.py plan   <loopdir>                       # produce a plan
  python3 loop_tick.py resolve <loopdir> --step S --result R [--metric K=V ...]

Why split into plan/resolve: the work itself runs OUTSIDE this Python process
(Apify, Supabase, the agent doing /synergy-discover etc.). We cleanly separate
"deciding what to do" from "recording what was done" so neither side has to
hold the other's state.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

# Local modules
import loop_state as ls
import loop_verify as lv

try:
    import openpyxl  # type: ignore
except ImportError:
    print(
        "ERROR: openpyxl is required (pip install openpyxl)", file=sys.stderr
    )
    sys.exit(30)


# --- Step names ------------------------------------------------------------

STEP_REFRESH_FINGERPRINT = "REFRESH_FINGERPRINT"
STEP_DISCOVER_AUTHOR = "DISCOVER_AUTHOR"
STEP_DISCOVER_CONTENT = "DISCOVER_CONTENT"
STEP_HARVEST_CITATIONS = "HARVEST_CITATIONS"
STEP_STAGE_DRAFTS = "STAGE_DRAFTS"
STEP_BOOK_CALENDAR = "BOOK_CALENDAR"
STEP_IDLE = "IDLE"

ALL_STEPS = {
    STEP_REFRESH_FINGERPRINT,
    STEP_DISCOVER_AUTHOR,
    STEP_DISCOVER_CONTENT,
    STEP_HARVEST_CITATIONS,
    STEP_STAGE_DRAFTS,
    STEP_BOOK_CALENDAR,
    STEP_IDLE,
}

# Map each step to the slash command the agent should run.
STEP_TO_SLASH = {
    STEP_REFRESH_FINGERPRINT: "/synergy-engine:synergy-fingerprint",
    STEP_DISCOVER_AUTHOR: "/synergy-engine:synergy-discover author",
    STEP_DISCOVER_CONTENT: "/synergy-engine:synergy-discover content",
    STEP_HARVEST_CITATIONS: "/synergy-engine:synergy-cite-harvest",
    STEP_STAGE_DRAFTS: "(internal) stage Queued -> Ready for review",
    STEP_BOOK_CALENDAR: "(internal) book Google Calendar block",
    STEP_IDLE: "(no action)",
}


# --- Step plan + result ----------------------------------------------------


@dataclass
class StepPlan:
    step: str
    slash_command: str
    reason: str
    cursor_in: dict[str, Any]
    # Manifest the slash command should read. Persisted to
    # .loop-state/next_step.json so the slash command can `cat` it.
    manifest: dict[str, Any]
    exit_code: int = ls.EXIT_CONTINUE

    def to_dict(self) -> dict[str, Any]:
        return {
            "step": self.step,
            "slash_command": self.slash_command,
            "reason": self.reason,
            "cursor_in": self.cursor_in,
            "manifest": self.manifest,
            "exit_code": self.exit_code,
        }


# --- Tracker helpers -------------------------------------------------------


def _tracker_counts(tracker_path: Path) -> dict[str, int]:
    """Cheap aggregate over the tracker. Used by step selection +
    handoff.drafts_ready_for_review."""
    counts = {
        "total": 0,
        "not_started": 0,
        "queued": 0,
        "ready_for_review": 0,
        "posted": 0,
        "pending_accept": 0,
        "accepted": 0,
        "engaged": 0,
        "parked": 0,
        "cold": 0,
        "author_center_queued": 0,
        "content_center_queued": 0,
    }
    if not tracker_path.exists():
        return counts
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True, read_only=True)
    if "Outreach Tracker" not in wb.sheetnames:
        return counts
    ws = wb["Outreach Tracker"]
    iter_rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(iter_rows)]
    except StopIteration:
        return counts
    if "Status" not in headers or "Persona" not in headers:
        return counts
    si = headers.index("Status")
    pi = headers.index("Persona")
    for row in iter_rows:
        if all(v is None or v == "" for v in row):
            continue
        counts["total"] += 1
        status = str(row[si] or "").strip()
        persona = str(row[pi] or "").strip().lower()
        key = status.lower().replace(" ", "_").replace("/", "_")
        if key in counts:
            counts[key] += 1
        if status in ("Not started", "Queued"):
            # Heuristic split: any persona containing 'citation' is citation-
            # center (a separate flow); otherwise it counts toward whichever
            # discovery center most recently put it there. We track per-center
            # progress separately via state.discovery counters; this is just
            # for human-readable status.
            if "citation" in persona:
                pass  # citation-center counts via citation block, not here
            else:
                counts["author_center_queued"] += 1
    return counts


# --- Source hashing (cheap, dependency-free) -------------------------------


def _hash_paths(paths: list[Path]) -> str:
    h = hashlib.sha256()
    for p in sorted(paths):
        try:
            st = p.stat()
            h.update(p.name.encode())
            h.update(str(st.st_size).encode())
            h.update(str(int(st.st_mtime)).encode())
        except FileNotFoundError:
            h.update(b"<missing>")
            h.update(p.name.encode())
    return h.hexdigest()[:16]


def _fingerprint_source_hash(loopdir: Path, tracker_path: Path) -> str:
    """Hash the things that, when changed, should invalidate the fingerprint:
    topic-fingerprint.md, the engine config, and the anchor paper (best-effort
    discovery — same dir as tracker)."""
    candidates = [
        tracker_path.parent / "topic-fingerprint.md",
        tracker_path.parent / "synergy-engine-config.md",
    ]
    # Also hash any *.md file next to the tracker that looks like an anchor.
    if tracker_path.parent.exists():
        for f in tracker_path.parent.glob("*anchor*.md"):
            candidates.append(f)
        for f in tracker_path.parent.glob("*paper*.md"):
            candidates.append(f)
    return _hash_paths(candidates)


def _bibliography_hash(tracker_path: Path) -> str:
    """Hash the bibliography file the citation center reads from. Best-effort
    discovery — look for *.bib next to the tracker."""
    if not tracker_path.parent.exists():
        return ""
    bibs = list(tracker_path.parent.glob("*.bib"))
    return _hash_paths(bibs)


# --- Step selection (the cursor) -------------------------------------------


def _select_step(state: ls.State, tracker_path: Path) -> StepPlan:
    """The dispatcher's brain. Picks one step deterministically based on
    cursors + tracker state. Order of checks below IS the priority."""

    counts = _tracker_counts(tracker_path)
    n = state.bounds.n_per_batch

    # Update handoff.drafts_ready_for_review from live tracker so check #6
    # has accurate input.
    state.handoff.drafts_ready_for_review = counts["ready_for_review"]

    # --- Goal-met check ---------------------------------------------------
    if (
        counts["ready_for_review"] >= n
        and state.handoff.last_calendar_block_id
    ):
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason=(
                f"goal met: {counts['ready_for_review']} drafts ready and "
                f"calendar block {state.handoff.last_calendar_block_id} booked"
            ),
            cursor_in={},
            manifest={},
            exit_code=ls.EXIT_GOAL_MET,
        )

    # --- Book calendar block if drafts are ready but no block yet --------
    if (
        counts["ready_for_review"] >= n
        and not state.handoff.last_calendar_block_id
    ):
        return StepPlan(
            step=STEP_BOOK_CALENDAR,
            slash_command=STEP_TO_SLASH[STEP_BOOK_CALENDAR],
            reason=f"{counts['ready_for_review']} drafts ready; no calendar block booked",
            cursor_in={"drafts_ready": counts["ready_for_review"]},
            manifest={
                "window": state.handoff.calendar_window,
                "title": state.handoff.calendar_title_pattern.format(
                    project=state.project, n=counts["ready_for_review"]
                ),
                "drafts_ready": counts["ready_for_review"],
            },
        )

    # --- Refresh fingerprint if sources changed --------------------------
    current_fp_hash = _fingerprint_source_hash(
        Path(state.tracker_path).parent.parent / ".loop-state",  # unused, kept for parity
        Path(state.tracker_path),
    )
    if (
        state.fingerprint.source_hash is None
        or state.fingerprint.source_hash != current_fp_hash
    ):
        return StepPlan(
            step=STEP_REFRESH_FINGERPRINT,
            slash_command=STEP_TO_SLASH[STEP_REFRESH_FINGERPRINT],
            reason=(
                "fingerprint never built"
                if state.fingerprint.source_hash is None
                else "fingerprint sources changed since last build"
            ),
            cursor_in={
                "last_built_utc": state.fingerprint.last_built_utc,
                "old_hash": state.fingerprint.source_hash,
                "new_hash": current_fp_hash,
            },
            manifest={
                "expected_outputs": ["topic-fingerprint.md"],
                "resolve_hash": current_fp_hash,
            },
        )

    # --- Stage drafts if we have enough queued rows ----------------------
    short = n - counts["ready_for_review"]
    if counts["queued"] + counts["not_started"] >= short and short > 0:
        return StepPlan(
            step=STEP_STAGE_DRAFTS,
            slash_command=STEP_TO_SLASH[STEP_STAGE_DRAFTS],
            reason=(
                f"need {short} more drafts to reach batch size {n}; "
                f"{counts['queued']} queued + {counts['not_started']} not_started available"
            ),
            cursor_in={
                "have_ready": counts["ready_for_review"],
                "have_queued": counts["queued"],
                "need": short,
            },
            manifest={
                "n_to_stage": short,
                "batch_target": n,
                "due_rule": (
                    "Status in {Not started, Queued} or Status in "
                    "{Posted, Accepted, Engaged} with Next Action Date <= today"
                ),
            },
        )

    # --- Harvest citations if bibliography changed -----------------------
    current_bib_hash = _bibliography_hash(Path(state.tracker_path))
    if current_bib_hash and (
        state.citation.bibliography_hash != current_bib_hash
        or state.citation.registry_pending_verify > 0
    ):
        return StepPlan(
            step=STEP_HARVEST_CITATIONS,
            slash_command=STEP_TO_SLASH[STEP_HARVEST_CITATIONS],
            reason=(
                "bibliography changed"
                if state.citation.bibliography_hash != current_bib_hash
                else f"{state.citation.registry_pending_verify} pending verifications"
            ),
            cursor_in={
                "old_hash": state.citation.bibliography_hash,
                "new_hash": current_bib_hash,
                "pending": state.citation.registry_pending_verify,
            },
            manifest={
                "resolve_hash": current_bib_hash,
                "respect_connect_envelope": True,
            },
        )

    # --- Discovery: alternate author <-> content centers -----------------
    last_author = state.discovery.author_center_cursor or ""
    last_content = state.discovery.content_center_cursor or ""
    # Pick whichever ran less recently. Empty cursor is oldest.
    if last_author <= last_content:
        return StepPlan(
            step=STEP_DISCOVER_AUTHOR,
            slash_command=STEP_TO_SLASH[STEP_DISCOVER_AUTHOR],
            reason=f"author-center queue thin ({counts['queued']} queued); cursor {last_author or '(never run)'}",
            cursor_in={
                "last_author_cursor": last_author,
                "queued_total": counts["queued"],
            },
            manifest={
                "center": "author",
                "exclude_publicidentifiers": True,
                "respect_cadence": True,
            },
        )
    return StepPlan(
        step=STEP_DISCOVER_CONTENT,
        slash_command=STEP_TO_SLASH[STEP_DISCOVER_CONTENT],
        reason=f"content-center queue thin; cursor {last_content or '(never run)'}",
        cursor_in={
            "last_content_cursor": last_content,
            "queued_total": counts["queued"],
        },
        manifest={
            "center": "content",
            "exclude_publicidentifiers": True,
            "respect_cadence": True,
        },
    )


# --- Bounds gates ----------------------------------------------------------


def _check_bounds_before_tick(state: ls.State) -> Optional[StepPlan]:
    """Returns a halt-flavored StepPlan if any hard bound is tripped."""

    if ls.is_halted(state):
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason=f"loop halted: {state.loop.halt_reason}",
            cursor_in={},
            manifest={},
            exit_code=ls.EXIT_HARD_STOP,
        )

    if state.bounds.budget_usd_remaining <= 0:
        ls.halt(state, "budget exhausted")
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason="hard stop: budget exhausted",
            cursor_in={},
            manifest={},
            exit_code=ls.EXIT_HARD_STOP,
        )

    if (
        state.bounds.wall_clock_deadline_utc
        and ls.utcnow_iso() > state.bounds.wall_clock_deadline_utc
    ):
        ls.halt(state, "wall-clock deadline passed")
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason="hard stop: wall-clock deadline passed",
            cursor_in={},
            manifest={},
            exit_code=ls.EXIT_HARD_STOP,
        )

    if (
        state.loop.consecutive_no_progress_ticks
        >= state.bounds.max_consecutive_no_progress
    ):
        ls.halt(state, f"no progress for {state.loop.consecutive_no_progress_ticks} ticks")
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason=f"hard stop: no progress for {state.loop.consecutive_no_progress_ticks} ticks",
            cursor_in={},
            manifest={},
            exit_code=ls.EXIT_HARD_STOP,
        )

    # Quota-based back-off: if today's calls already exhausted any per-day
    # ceiling, signal quota and skip work.
    if (
        state.quotas.apify_calls_today >= state.bounds.max_apify_per_day
        or state.quotas.apollo_calls_today >= state.bounds.max_apollo_per_day
        or state.quotas.openalex_calls_today >= state.bounds.max_openalex_per_day
    ):
        return StepPlan(
            step=STEP_IDLE,
            slash_command=STEP_TO_SLASH[STEP_IDLE],
            reason="quota ceiling hit for today; backing off until tomorrow",
            cursor_in={
                "apify": state.quotas.apify_calls_today,
                "apollo": state.quotas.apollo_calls_today,
                "openalex": state.quotas.openalex_calls_today,
            },
            manifest={},
            exit_code=ls.EXIT_QUOTA,
        )

    return None


# --- Persist plan / load plan ---------------------------------------------


def _write_next_step(loopdir: Path, plan: StepPlan) -> None:
    """The slash command reads this to know what to do."""
    (loopdir / "next_step.json").write_text(
        json.dumps(plan.to_dict(), indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _read_next_step(loopdir: Path) -> Optional[dict[str, Any]]:
    p = loopdir / "next_step.json"
    if not p.exists():
        return None
    return json.loads(p.read_text(encoding="utf-8"))


# --- Plan / resolve drivers -----------------------------------------------


def cmd_plan(loopdir: Path) -> int:
    """Decide what this tick should do. Writes next_step.json and prints
    the plan as JSON on stdout."""
    try:
        state = ls.load(loopdir)
    except ls.StateNotFoundError as e:
        print(f"ERROR: no state.json at {e}; run /synergy-loop-start first", file=sys.stderr)
        return ls.EXIT_TRANSIENT
    except ls.StateSchemaError as e:
        print(f"ERROR: state.json schema invalid: {e}", file=sys.stderr)
        return ls.EXIT_TRANSIENT

    # Daily counter roll
    rolled = ls.roll_daily_counters_if_needed(state)
    if rolled:
        ls.save(loopdir, state)

    halt_plan = _check_bounds_before_tick(state)
    if halt_plan:
        ls.save(loopdir, state)  # persist any halt reason set in the gate
        _write_next_step(loopdir, halt_plan)
        print(json.dumps(halt_plan.to_dict(), indent=2, sort_keys=True))
        return halt_plan.exit_code

    tracker_path = Path(state.tracker_path)
    plan = _select_step(state, tracker_path)
    ls.save(loopdir, state)  # _select_step may have updated drafts_ready_for_review

    # --- Open the chain for this tick: mint tick_id + plan_nonce ---------
    # IDLE/halt plans don't open a chain (no downstream work to gate).
    if plan.step != STEP_IDLE and plan.exit_code == ls.EXIT_CONTINUE:
        state, tick_id, plan_nonce = ls.start_tick(loopdir)
        plan.manifest = dict(plan.manifest)  # copy before mutating
        plan.manifest["tick_id"] = tick_id
        plan.manifest["plan_nonce"] = plan_nonce
        plan.manifest["release_owner"] = state.release_owner
        plan.manifest["chain_audit_path"] = str(loopdir / "chain-audit.jsonl")

    _write_next_step(loopdir, plan)
    print(json.dumps(plan.to_dict(), indent=2, sort_keys=True))
    return plan.exit_code


def cmd_resolve(
    loopdir: Path,
    step: str,
    result: str,
    metrics: dict[str, str],
    rows_written: list[str],
    fingerprint_refreshed: bool,
    plan_nonce: str,
) -> int:
    """Record the outcome of the slash command the agent just ran. Updates
    cursors, runs the verifier, persists state, exits with the right code.

    plan_nonce is REQUIRED and must match the active chain's plan_nonce. The
    only exception is IDLE resolves (no chain was opened), which are allowed
    to pass plan_nonce="".
    """
    if step not in ALL_STEPS:
        print(f"ERROR: unknown step {step!r}", file=sys.stderr)
        return ls.EXIT_TRANSIENT
    if result not in ("ok", "fail", "partial", "skip"):
        print(f"ERROR: result must be ok|fail|partial|skip", file=sys.stderr)
        return ls.EXIT_TRANSIENT

    try:
        state_before = ls.load(loopdir)
    except ls.StateNotFoundError as e:
        print(f"ERROR: no state.json at {e}", file=sys.stderr)
        return ls.EXIT_TRANSIENT

    state = ls.load(loopdir)  # working copy to mutate
    plan_d = _read_next_step(loopdir) or {}

    # --- Chain gate: refuse to resolve without the active plan_nonce ------
    # IDLE has no chain to gate.
    if step != STEP_IDLE:
        chain = state_before.chain
        active_plan = chain.plan_nonce if chain else None
        if not active_plan:
            print(
                "ERROR: no open chain for this tick; cannot resolve. Run plan first.",
                file=sys.stderr,
            )
            return ls.EXIT_TRANSIENT
        if not plan_nonce:
            print(
                "ERROR: --plan-nonce is required for resolve (non-IDLE step)",
                file=sys.stderr,
            )
            return ls.EXIT_TRANSIENT
        if plan_nonce != active_plan:
            print(
                f"ERROR: plan_nonce mismatch; refusing to resolve. "
                f"(active chain tick_id={chain.tick_id})",
                file=sys.stderr,
            )
            # Audit the rejection by attempting a consume — it will record
            # the mismatch in chain-audit.jsonl.
            try:
                ls.consume_nonce(loopdir, ls.NONCE_PLAN, plan_nonce)
            except ls.NonceError:
                pass
            return ls.EXIT_HARD_STOP

    # --- Apply cursor advances based on step + result ---------------------

    if result == "fail":
        # Don't advance cursors on failure; verifier will note no-progress.
        pass
    elif step == STEP_REFRESH_FINGERPRINT:
        manifest = plan_d.get("manifest", {})
        new_hash = manifest.get("resolve_hash")
        if new_hash:
            state.fingerprint.source_hash = new_hash
        state.fingerprint.last_built_utc = ls.utcnow_iso()
    elif step == STEP_DISCOVER_AUTHOR:
        state.discovery.author_center_cursor = ls.utcnow_iso()
        state.discovery.last_discover_utc = ls.utcnow_iso()
        delta = int(metrics.get("queued_added", "0") or "0")
        state.discovery.queued_since_last_run += delta
        state.quotas.apify_calls_today += int(metrics.get("apify_calls", "0") or "0")
    elif step == STEP_DISCOVER_CONTENT:
        state.discovery.content_center_cursor = ls.utcnow_iso()
        state.discovery.last_discover_utc = ls.utcnow_iso()
        delta = int(metrics.get("queued_added", "0") or "0")
        state.discovery.queued_since_last_run += delta
        state.quotas.apify_calls_today += int(metrics.get("apify_calls", "0") or "0")
    elif step == STEP_HARVEST_CITATIONS:
        manifest = plan_d.get("manifest", {})
        new_hash = manifest.get("resolve_hash")
        if new_hash:
            state.citation.bibliography_hash = new_hash
        state.citation.last_harvest_utc = ls.utcnow_iso()
        state.citation.registry_pending_verify = int(
            metrics.get("pending_after", "0") or "0"
        )
        state.quotas.apollo_calls_today += int(metrics.get("apollo_calls", "0") or "0")
        state.quotas.openalex_calls_today += int(
            metrics.get("openalex_calls", "0") or "0"
        )
    elif step == STEP_STAGE_DRAFTS:
        # No cursor advance; the tracker IS the source of truth. We'll re-read
        # counts after the verifier.
        pass
    elif step == STEP_BOOK_CALENDAR:
        block_id = metrics.get("calendar_block_id")
        if block_id:
            state.handoff.last_calendar_block_id = block_id
            state.handoff.last_calendar_block_utc = metrics.get(
                "calendar_block_utc", ls.utcnow_iso()
            )
    elif step == STEP_IDLE:
        pass  # nothing to advance

    # Budget burn (caller may report --metric budget_usd_spent=0.42)
    spent = float(metrics.get("budget_usd_spent", "0") or "0")
    if spent > 0:
        state.bounds.budget_usd_remaining = max(
            0.0, state.bounds.budget_usd_remaining - spent
        )

    # Persist mid-stream so the verifier sees the post-step state.
    ls.save(loopdir, state)

    # --- Verifier ---------------------------------------------------------
    rows_written_keys = {r for r in rows_written if r}
    report = lv.run_all_checks(
        loop_state_dir=loopdir,
        tracker_path=Path(state.tracker_path),
        rows_written_keys=rows_written_keys,
        state_before=state_before,
        fingerprint_refreshed=(step == STEP_REFRESH_FINGERPRINT and result == "ok"),
        step_just_ran=step,
        plan_nonce=(plan_nonce if step != STEP_IDLE else None),
    )

    # The verifier may have set halt_reason (check #4). Reload to capture it.
    state = ls.load(loopdir)

    # --- Chain close: verify the verifier echoed our plan_nonce verbatim --
    # and consume the verifier's own nonce. Any failure is a HARD STOP — a
    # passing-shaped report that didn't actually do the work cannot close
    # the chain.
    if step != STEP_IDLE:
        if report.plan_nonce_echoed != plan_nonce:
            ls.halt(
                state,
                f"verifier did not echo plan_nonce (got {report.plan_nonce_echoed!r})",
            )
            ls.clear_chain(loopdir, reason="verifier_echo_failed")
            ls.save(loopdir, state)
            print(
                "ERROR: verifier did not echo plan_nonce; chain rejected",
                file=sys.stderr,
            )
            return ls.EXIT_HARD_STOP
        if not report.verifier_nonce:
            ls.halt(state, "verifier produced no verifier_nonce")
            ls.clear_chain(loopdir, reason="missing_verifier_nonce")
            ls.save(loopdir, state)
            print("ERROR: verifier produced no verifier_nonce", file=sys.stderr)
            return ls.EXIT_HARD_STOP
        try:
            ls.consume_nonce(loopdir, ls.NONCE_VERIFIER, report.verifier_nonce)
        except ls.NonceError as e:
            ls.halt(state, f"verifier_nonce consume failed: {e}")
            ls.clear_chain(loopdir, reason=f"verifier_nonce_consume_failed:{e}")
            ls.save(loopdir, state)
            print(f"ERROR: verifier_nonce consume failed: {e}", file=sys.stderr)
            return ls.EXIT_HARD_STOP

    # --- Tick bookkeeping -------------------------------------------------
    made_progress = report.made_progress
    exit_code: int
    if report.hard_stop_reason:
        exit_code = ls.EXIT_HARD_STOP
        status_str = f"hard_stop: {report.hard_stop_reason}"
    elif result == "fail":
        exit_code = ls.EXIT_TRANSIENT
        status_str = f"step={step} failed; will retry next tick"
    elif not report.all_passed:
        exit_code = ls.EXIT_TRANSIENT
        status_str = f"step={step} ok but verifier soft-failed"
    elif (
        state.handoff.drafts_ready_for_review >= state.bounds.n_per_batch
        and state.handoff.last_calendar_block_id
    ):
        exit_code = ls.EXIT_GOAL_MET
        status_str = "goal met: batch ready + calendar booked"
    else:
        exit_code = ls.EXIT_CONTINUE
        status_str = f"step={step} ok"

    ls.record_tick(state, status=status_str, exit_code=exit_code, made_progress=made_progress)
    ls.save(loopdir, state)

    # --- Per-tick log -----------------------------------------------------
    log_dir = loopdir
    log_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    (log_dir / f"tick-{ts}.log").write_text(
        json.dumps(
            {
                "tick_utc": ls.utcnow_iso(),
                "step": step,
                "result": result,
                "metrics": metrics,
                "rows_written": sorted(rows_written_keys),
                "verifier": report.to_dict(),
                "exit_code": exit_code,
                "status": status_str,
            },
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    # --- User-facing single-line summary ---------------------------------
    print(
        json.dumps(
            {
                "tick_utc": ls.utcnow_iso(),
                "step": step,
                "result": result,
                "verifier_ok": report.all_passed,
                "hard_stop": report.hard_stop_reason,
                "ticks_total": state.loop.ticks_total,
                "drafts_ready": state.handoff.drafts_ready_for_review,
                "exit_code": exit_code,
                "status": status_str,
            },
            indent=2,
            sort_keys=True,
        )
    )

    # --- Chain close: burn the plan_nonce and clear the chain --------------
    # Even on transient/soft-fail we close the chain — the chain represents
    # one tick, and the next tick will mint a fresh one. The exit code tells
    # the caller whether to retry; the chain itself is one-shot.
    if step != STEP_IDLE:
        try:
            ls.consume_nonce(loopdir, ls.NONCE_PLAN, plan_nonce)
        except ls.NonceError as e:
            # The plan_nonce we already gate-checked above; a failure here
            # would mean state was mutated under us. Log and continue — the
            # audit row is already written.
            print(f"WARN: plan_nonce consume returned {e}", file=sys.stderr)
        ls.clear_chain(
            loopdir,
            reason=f"tick_resolved:exit={exit_code}:step={step}:result={result}",
        )

    # Clean up the next_step.json now that it's resolved.
    nsp = loopdir / "next_step.json"
    if nsp.exists():
        nsp.unlink()

    return exit_code


# --- Mid-tick chain helpers ------------------------------------------------


def cmd_issue_nonce(loopdir: Path, kind: str, plan_nonce: str) -> int:
    """Mint a tracker_nonce or calendar_nonce against the currently open
    chain. The step-executor agent calls this RIGHT BEFORE the tracker write
    or the calendar create, presents the returned nonce to the writer, then
    the writer (or its wrapper) calls cmd_consume_nonce to burn it.

    Prints {"kind": ..., "nonce": ...} as JSON on stdout.
    """
    if kind not in (ls.NONCE_TRACKER, ls.NONCE_CALENDAR):
        print(
            f"ERROR: kind must be {ls.NONCE_TRACKER!r} or {ls.NONCE_CALENDAR!r}, got {kind!r}",
            file=sys.stderr,
        )
        return ls.EXIT_TRANSIENT
    try:
        nonce = ls.issue_nonce(loopdir, kind, plan_nonce)
    except ls.NonceError as e:
        print(f"ERROR: issue_nonce failed: {e}", file=sys.stderr)
        return ls.EXIT_HARD_STOP
    print(json.dumps({"kind": kind, "nonce": nonce}, indent=2, sort_keys=True))
    return ls.EXIT_CONTINUE


def cmd_consume_nonce(loopdir: Path, kind: str, presented: str) -> int:
    """Burn a previously-issued tracker_nonce or calendar_nonce. The writer
    wrapper calls this AFTER the side effect (tracker row write / calendar
    event create) has succeeded. Mismatch or replay raises NonceError, which
    closes the chain with a HARD STOP — the audit row is already written.

    Prints {"kind": ..., "consumed": true} on success.
    """
    if kind not in (ls.NONCE_TRACKER, ls.NONCE_CALENDAR):
        print(
            f"ERROR: kind must be {ls.NONCE_TRACKER!r} or {ls.NONCE_CALENDAR!r}, got {kind!r}",
            file=sys.stderr,
        )
        return ls.EXIT_TRANSIENT
    try:
        ls.consume_nonce(loopdir, kind, presented)
    except ls.NonceError as e:
        print(f"ERROR: consume_nonce failed: {e}", file=sys.stderr)
        return ls.EXIT_HARD_STOP
    print(
        json.dumps(
            {"kind": kind, "consumed": True}, indent=2, sort_keys=True
        )
    )
    return ls.EXIT_CONTINUE


# --- CLI -------------------------------------------------------------------


def _parse_metrics(raw: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for kv in raw:
        if "=" not in kv:
            raise SystemExit(f"metric must be KEY=VALUE, got {kv!r}")
        k, v = kv.split("=", 1)
        out[k.strip()] = v.strip()
    return out


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    sub = p.add_subparsers(dest="cmd", required=True)

    sp_plan = sub.add_parser("plan", help="Decide what to do this tick.")
    sp_plan.add_argument("loopdir", type=Path)

    sp_res = sub.add_parser("resolve", help="Record what was done this tick.")
    sp_res.add_argument("loopdir", type=Path)
    sp_res.add_argument("--step", required=True, choices=sorted(ALL_STEPS))
    sp_res.add_argument(
        "--result", required=True, choices=("ok", "fail", "partial", "skip")
    )
    sp_res.add_argument(
        "--metric",
        action="append",
        default=[],
        help="Repeatable KEY=VALUE. Common keys: queued_added, apify_calls, apollo_calls, openalex_calls, pending_after, calendar_block_id, calendar_block_utc, budget_usd_spent",
    )
    sp_res.add_argument(
        "--rows-written",
        default="",
        help="Comma-separated Target names this step wrote to the tracker.",
    )
    sp_res.add_argument(
        "--fingerprint-refreshed",
        action="store_true",
        help="Set when the resolved step refreshed the fingerprint.",
    )
    sp_res.add_argument(
        "--plan-nonce",
        default="",
        help=(
            "REQUIRED for non-IDLE steps. The plan_nonce minted at plan time "
            "and persisted in next_step.json -> manifest.plan_nonce. The "
            "resolve refuses to run if this doesn't match the active chain."
        ),
    )

    sp_itn = sub.add_parser(
        "issue-tracker-nonce",
        help="Mint a tracker_nonce against the active chain before a tracker write.",
    )
    sp_itn.add_argument("loopdir", type=Path)
    sp_itn.add_argument("--plan-nonce", required=True)

    sp_icn = sub.add_parser(
        "issue-calendar-nonce",
        help="Mint a calendar_nonce against the active chain before a calendar create.",
    )
    sp_icn.add_argument("loopdir", type=Path)
    sp_icn.add_argument("--plan-nonce", required=True)

    sp_cn = sub.add_parser(
        "consume-nonce",
        help="Burn a previously-issued tracker_nonce or calendar_nonce.",
    )
    sp_cn.add_argument("loopdir", type=Path)
    sp_cn.add_argument("--kind", required=True, choices=("tracker", "calendar"))
    sp_cn.add_argument("--nonce", required=True)

    args = p.parse_args(argv)

    if args.cmd == "plan":
        return cmd_plan(args.loopdir)
    elif args.cmd == "resolve":
        rows = [r.strip() for r in args.rows_written.split(",") if r.strip()]
        metrics = _parse_metrics(args.metric)
        return cmd_resolve(
            loopdir=args.loopdir,
            step=args.step,
            result=args.result,
            metrics=metrics,
            rows_written=rows,
            fingerprint_refreshed=args.fingerprint_refreshed,
            plan_nonce=args.plan_nonce,
        )
    elif args.cmd == "issue-tracker-nonce":
        return cmd_issue_nonce(args.loopdir, ls.NONCE_TRACKER, args.plan_nonce)
    elif args.cmd == "issue-calendar-nonce":
        return cmd_issue_nonce(args.loopdir, ls.NONCE_CALENDAR, args.plan_nonce)
    elif args.cmd == "consume-nonce":
        return cmd_consume_nonce(args.loopdir, args.kind, args.nonce)
    return ls.EXIT_TRANSIENT


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
