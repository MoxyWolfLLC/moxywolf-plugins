#!/usr/bin/env python3
"""Create the synergy-engine outreach tracker xlsx (Outreach Tracker + How this works).

Idempotent guard: refuses to overwrite an existing file. Usage:
    python3 tracker_init.py --out "/path/to/<name>-linkedin-outreach-tracker.xlsx"
Requires openpyxl (pip install openpyxl --break-system-packages).
"""
import argparse, os, sys

HEADERS = ["Target","Persona","Tier","Path","Synergy","LinkedIn Profile","Last Touch",
           "Liked","Commented","Comment / engagement summary","Cited URL",
           "Connect / DM","Status","Next Action","Next Action Date","Notes"]

LEGEND = [
 "Synergy Engine — Outreach Tracker",
 "",
 "One row per target. This sheet is the engine's memory, dedupe, and queue across both topic centers.",
 "",
 "PATHS (see topic-synergy-methodology.md):",
 "  A = their post maps to one of your /answers categories -> cite their work in your answer, like, comment WITH link, then connect/DM.",
 "  B = relevant but no category fit -> like, comment that engages/challenges/poses a question, link sparingly.",
 "  GA = anchor-paper cite -> cite your paper/POV in the comment (content center).",
 "  cite-only = they don't post -> cite their work in your content, then InMail.",
 "",
 "STATUS: Not started | Queued | Ready for review | Posted | Pending accept | Accepted | Replied | Engaged | Parked | Cold",
 "",
 "DUE rule: Status in {Not started, Queued, Ready for review}, OR Status in {Posted, Accepted, Engaged} with Next Action Date <= today. Skip Parked.",
 "",
 "CADENCE: comment same day; connect/DM 2-3 days later; <=5 fresh targets per run; don't re-comment within ~3 of a person's posts; re-sweep each cycle.",
 "",
 "DEDUPE: before discovery, pass every publicIdentifier already here as the exclude-list to scoring.",
 "",
 "Synergy colors: green=High, yellow=Medium, orange=Low, grey=Cold/Parked.",
]

def build(out):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Outreach Tracker"
    hfill = PatternFill("solid", fgColor="1F3864"); hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin,right=thin,top=thin,bottom=thin)
    for c,h in enumerate(HEADERS,1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill=hfill; cell.font=hfont; cell.alignment=center; cell.border=border
    widths=[16,12,10,10,14,34,12,7,14,38,40,22,16,40,15,40]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=30
    ws2 = wb.create_sheet("How this works")
    for i,line in enumerate(LEGEND,1):
        cell = ws2.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        if i==1: cell.font = Font(bold=True, size=13)
        elif line.endswith(":") or line[:6] in ("PATHS ","STATUS","CADENC","DEDUPE","DUE ru","Synerg"): cell.font = Font(bold=True, size=10)
        else: cell.font = Font(size=10)
    ws2.column_dimensions["A"].width = 120
    wb.save(out)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    a = ap.parse_args()
    if os.path.exists(a.out):
        print(f"REFUSING to overwrite existing tracker: {a.out}"); sys.exit(1)
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    build(a.out)
    print(f"created tracker: {a.out}")

if __name__ == "__main__":
    main()
