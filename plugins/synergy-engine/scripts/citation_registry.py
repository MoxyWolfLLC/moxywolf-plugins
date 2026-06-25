#!/usr/bin/env python3
"""Create or populate the synergy-engine citation registry xlsx.

Three sheets: "Cited Authors" (one row per reachable person), "Non-targets"
(bibliography entries that aren't reachable people), and "How this works" (legend).
Schema in references/citation-registry-schema.md.

Two modes:
  --init   create an empty registry (headers + legend only). Idempotent: refuses
           to overwrite an existing file.
  --from-json <authors.json> [--nontargets-json <nt.json>] [--force]
           build/overwrite a registry from harvested JSON. Each author object may
           carry any of the column keys (see AUTHOR_KEYS); missing keys -> blank.

Usage:
  python3 citation_registry.py --init --out "/path/<paper>-citation-registry.xlsx"
  python3 citation_registry.py --from-json authors.json --out "/path/reg.xlsx" [--nontargets-json nt.json] [--force]

Requires openpyxl (pip install openpyxl --break-system-packages).
"""
import argparse, json, os, sys

AUTHORS_HEADERS = [
    "Author", "Priority", "ORCID", "Cited Affiliation", "Cited Works", "Sections",
    "How Used", "Apollo Title", "Current Org", "Org Domain", "Email", "Email Status",
    "Verify", "Moved", "LinkedIn URL", "LinkedIn Conf", "Email Sent", "Connect Sent",
    "Status", "Next Action", "Draft Email", "Draft Connect Note",
]
# JSON key per column (snake/loose -> exact). Accepts a few aliases.
AUTHOR_KEYS = [
    "author", "priority", "orcid", "cited_affiliation", "cited_works", "sections",
    "how_used", "apollo_title", "current_org", "org_domain", "email", "email_status",
    "verify", "moved", "linkedin_url", "linkedin_conf", "email_sent", "connect_sent",
    "status", "next_action", "draft_email", "draft_connect_note",
]
ALIASES = {
    "author": ["name"], "current_org": ["apollo_org", "org"], "linkedin_url": ["url", "linkedin"],
    "linkedin_conf": ["conf"], "cited_works": ["works"], "draft_connect_note": ["note", "connect_note"],
    "draft_email": ["email_body"],
}
NONTARGET_HEADERS = ["Reference", "Type", "Note"]

LEGEND = [
    "Synergy Engine - Citation Registry",
    "",
    "One row per reachable cited author (sheet 'Cited Authors'). The citation center's memory, dedupe, and queue.",
    "",
    "PRIORITY: load-bearing (central to a claim) | primary author | co-author.",
    "",
    "VERIFY (the gate): ok | REVIEW | FALSE POSITIVE - exclude. Every name+org Apollo match must pass before outreach;",
    "  LinkedIn-anchored matches pass automatically. FALSE POSITIVE rows are excluded from sending.",
    "",
    "LinkedIn Conf: verified (Apify) | REVIEW (Apify) | WRONG (Apify) - namesake | unverified (private profile) | sent/confirmed.",
    "  Only 'verified' (or hand-confirmed) URLs go into the send queue.",
    "",
    "STATUS: Not started | Drafted | Email sent | Connect pending | Accepted | Replied | Engaged | Excluded | Held.",
    "",
    "EMAIL structure: their bibliographic reference first, then how we used it, then our paper ONCE. BCC dorianc@moxywolf.com.",
    "CONNECT note: hook-free, <=300 chars, acknowledge the citation and connect. No promise.",
    "",
    "SEND discipline (references/outreach-channels.md): type the note with separate click+type calls, zoom-verify the",
    "  start AND end before Send; 2nd-degree=direct Connect, 3rd=under More, high-follower=Follow+More; confirm 'Pending';",
    "  never withdraw to re-send (3-week lockout); some profiles email-gate the connect (use the enriched email we hold).",
    "",
    "DEDUPE: ORCID first, normalized name fallback. Pass existing ORCIDs/names as the exclude-list before a re-harvest.",
    "",
    "Non-targets sheet: bibliography entries that aren't reachable people (organization | legal case | statute | classic-unreachable).",
]


def _get(obj, key):
    if key in obj and obj[key] not in (None, ""):
        return obj[key]
    for alt in ALIASES.get(key, []):
        if alt in obj and obj[alt] not in (None, ""):
            return obj[alt]
    return ""


def _style_header(ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hfill = PatternFill("solid", fgColor="1F3864")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hfill; cell.font = hfont; cell.alignment = center; cell.border = border
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30


def build(out, authors=None, nontargets=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Cited Authors"
    _style_header(ws, AUTHORS_HEADERS)
    widths = [20, 14, 20, 22, 34, 18, 34, 22, 24, 18, 26, 14, 18, 26, 34, 22, 12, 12, 16, 30, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    grey = PatternFill("solid", fgColor="D9D9D9")
    green = PatternFill("solid", fgColor="C6EFCE")
    orange = PatternFill("solid", fgColor="FCE4D6")
    bold = Font(bold=True)
    for r, a in enumerate(authors or [], start=2):
        for c, key in enumerate(AUTHOR_KEYS, 1):
            v = _get(a, key)
            if isinstance(v, list):
                v = "; ".join(str(x) for x in v)
            ws.cell(row=r, column=c, value=v)
        verify = str(_get(a, "verify")).upper()
        status = str(_get(a, "status")).lower()
        conf = str(_get(a, "linkedin_conf")).lower()
        if "false positive" in verify or status == "excluded":
            for c in range(1, len(AUTHORS_HEADERS) + 1):
                ws.cell(row=r, column=c).fill = grey
        if str(_get(a, "priority")).lower().startswith("load"):
            ws.cell(row=r, column=1).font = bold
        if conf.startswith("verified"):
            ws.cell(row=r, column=16).fill = green
        elif conf.startswith("wrong") or conf.startswith("review"):
            ws.cell(row=r, column=16).fill = orange

    ws2 = wb.create_sheet("Non-targets")
    _style_header(ws2, NONTARGET_HEADERS)
    for i, w in enumerate([60, 22, 50], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r, nt in enumerate(nontargets or [], start=2):
        ws2.cell(row=r, column=1, value=_get(nt, "reference"))
        ws2.cell(row=r, column=2, value=_get(nt, "type"))
        ws2.cell(row=r, column=3, value=_get(nt, "note"))

    ws3 = wb.create_sheet("How this works")
    for i, line in enumerate(LEGEND, 1):
        cell = ws3.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(horizontal="left", vertical="top")
        if i == 1:
            cell.font = Font(bold=True, size=13)
        elif line.endswith(":") or line[:6] in ("PRIORI", "VERIFY", "STATUS", "EMAIL ", "CONNEC", "SEND d", "DEDUPE", "Linked", "Non-ta"):
            cell.font = Font(bold=True, size=10)
        else:
            cell.font = Font(size=10)
    ws3.column_dimensions["A"].width = 120
    wb.save(out)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", required=True)
    ap.add_argument("--init", action="store_true", help="create an empty registry (refuses to overwrite)")
    ap.add_argument("--from-json", help="authors JSON (list of objects)")
    ap.add_argument("--nontargets-json", help="non-targets JSON (list of objects)")
    ap.add_argument("--force", action="store_true", help="allow overwrite in --from-json mode")
    a = ap.parse_args()

    if a.init and os.path.exists(a.out):
        print(f"REFUSING to overwrite existing registry: {a.out}"); sys.exit(1)
    if a.from_json and os.path.exists(a.out) and not a.force:
        print(f"REFUSING to overwrite without --force: {a.out}"); sys.exit(1)

    authors = json.load(open(a.from_json, encoding="utf-8")) if a.from_json else []
    nontargets = json.load(open(a.nontargets_json, encoding="utf-8")) if a.nontargets_json else []
    if isinstance(authors, dict):  # accept {name: {...}} maps
        authors = [{**v, "author": v.get("author", k)} for k, v in authors.items()]

    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    build(a.out, authors, nontargets)
    print(f"wrote registry: {a.out}  (authors={len(authors)}, non-targets={len(nontargets)})")


if __name__ == "__main__":
    main()
