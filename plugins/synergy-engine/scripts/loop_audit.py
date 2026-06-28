"""
loop_audit.py — override-rate + chain-of-custody audit roll-up for the
Synergy Engine prep loop.

This is the third leg of the Governed Autonomy stool. The paper (§10) warns
that a reviewer who just rubber-stamps everything is worse than no reviewer
at all, because the system now CLAIMS oversight while having none. The cure
is to measure the override rate — what fraction of the AI's drafts the human
actually changed, killed, or sent unedited — and surface it side-by-side
with the engineering-health metrics (verifier hard-stops, halt fires,
mid-tick interventions). Two numbers, two failure modes, one report.

Inputs
------
- The .loop-state/<project>/ directory:
    - state.json           — current snapshot (release_owner, halt_reason)
    - tick-*.log           — one JSON-line file per resolved tick
    - chain-audit.jsonl    — append-only nonce events (this is the
                              tamper-evident bit; the engineering metrics
                              come from here)
- The live tracker .xlsx (Outreach Tracker sheet) — read-only.
  This is where the GOVERNANCE metrics come from: a row whose Status moved
  Ready-for-review -> Posted IS a draft the human approved (or edited and
  approved). A row that ended up Parked or Cold IS a kill.

What the report contains
------------------------
Two top-level sections, both required (§§10–11 of the paper):

1. engineering_health — was the loop's own machinery exercised?
     - ticks_total
     - verifier_hard_stops      (count + rate per tick)
     - halt_fires               (count + per-tick rate)
     - mid_loop_interventions   (chain rejects, replays, mismatches)
     - chain_consume_failures   (the smoke gun for fake reviews)

2. governance_health — was the human actually overriding?
     - batches_sent             (rows with Status in {Posted, Pending Accept, Accepted, Engaged})
     - batches_staged           (rows with Status == Ready for review)
     - batches_killed           (rows with Status in {Parked, Cold})
     - sent_as_staged_ratio     (proxy for rubber-stamping; high = bad)
     - mean_time_to_first_send  (calendar_block_utc -> earliest Posted)
     - last_calendar_block_utc

The numbers are intentionally raw. The report does not say "this is bad" —
the paper is explicit that the threshold for "rubber-stamping" is a
judgment call that depends on the topic and the reviewer. We surface the
ratios; the human sets the line.

CLI
---
    python3 loop_audit.py <loopdir> <tracker.xlsx>           # pretty text
    python3 loop_audit.py <loopdir> <tracker.xlsx> --json    # machine

Exit code is always 0 unless the inputs themselves can't be read. The audit
is a READ-ONLY operation — it never mutates state, never opens the chain,
never appends to the audit log. It is safe to run from anywhere at any time.
"""

from __future__ import annotations

import argparse
import json
import statistics
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl  # type: ignore
except ImportError:
    print("ERROR: openpyxl is required (pip install openpyxl)", file=sys.stderr)
    sys.exit(30)

import loop_state as ls


# --- Status buckets --------------------------------------------------------
# Mirrors the tracker's Status vocabulary. Keep aligned with loop_verify and
# normalize_rows.

SENT_STATUSES = {"Posted", "Pending Accept", "Accepted", "Engaged"}
STAGED_STATUSES = {"Ready for review"}
KILLED_STATUSES = {"Parked", "Cold"}
QUEUED_STATUSES = {"Not started", "Queued"}


# --- Tick log reader -------------------------------------------------------


def _read_tick_logs(loopdir: Path) -> list[dict[str, Any]]:
    """Each tick writes one tick-<utc>.log. Return them sorted by filename
    (filename embeds the UTC timestamp so lexicographic == chronological)."""
    if not loopdir.exists():
        return []
    out = []
    for f in sorted(loopdir.glob("tick-*.log")):
        try:
            out.append(json.loads(f.read_text(encoding="utf-8")))
        except (json.JSONDecodeError, OSError):
            continue
    return out


def _read_audit_log(loopdir: Path) -> list[dict[str, Any]]:
    """The chain-audit.jsonl is append-only, one JSON object per line."""
    p = loopdir / "chain-audit.jsonl"
    if not p.exists():
        return []
    out = []
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            out.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return out


# --- Tracker reader --------------------------------------------------------


def _read_tracker_rows(tracker_path: Path) -> list[dict[str, Any]]:
    """Return a list of {Status, Persona, Target Name, Posted At, ...} dicts.
    Returns [] if the tracker can't be opened (e.g. missing sheet)."""
    if not tracker_path.exists():
        return []
    wb = openpyxl.load_workbook(str(tracker_path), data_only=True, read_only=True)
    if "Outreach Tracker" not in wb.sheetnames:
        return []
    ws = wb["Outreach Tracker"]
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(it)]
    except StopIteration:
        return []
    rows = []
    for row in it:
        if all(v is None or v == "" for v in row):
            continue
        rows.append({h: row[i] for i, h in enumerate(headers) if i < len(row)})
    return rows


# --- Engineering health ---------------------------------------------------


def _engineering_health(
    state: ls.State, ticks: list[dict[str, Any]], audit: list[dict[str, Any]]
) -> dict[str, Any]:
    total = len(ticks)
    hard_stops = sum(1 for t in ticks if t.get("exit_code") == ls.EXIT_HARD_STOP)
    transient = sum(1 for t in ticks if t.get("exit_code") == ls.EXIT_TRANSIENT)
    goal_met = sum(1 for t in ticks if t.get("exit_code") == ls.EXIT_GOAL_MET)

    # halt_fires: ticks whose status string starts with "hard_stop:". This
    # is a SUPERSET of the state's current halt_reason — a halt that was
    # later cleared and resumed still shows up here, which is what we want.
    halt_fires = sum(
        1 for t in ticks if str(t.get("status", "")).startswith("hard_stop")
    )

    # mid_loop_interventions: chain events that signal something went sideways
    # mid-tick. These are the events that would tell you a reviewer was
    # trying to short-circuit the system (or that the system caught it).
    intervention_events = {
        "nonce_consume_mismatch",
        "nonce_replayed",
        "nonce_missing",
        "verifier_echo_failed",
        "chain_overwritten",
    }
    interventions = [a for a in audit if a.get("event") in intervention_events]

    # chain_consume_failures: a strict subset of the above — these are the
    # ones that map most directly to "someone tried to close a chain without
    # doing the work."
    consume_fail_events = {"nonce_consume_mismatch", "nonce_replayed"}
    consume_fails = [a for a in audit if a.get("event") in consume_fail_events]

    return {
        "ticks_total": total,
        "verifier_hard_stops": hard_stops,
        "verifier_hard_stop_rate": round(hard_stops / total, 3) if total else 0.0,
        "transient_retries": transient,
        "goal_met_count": goal_met,
        "halt_fires": halt_fires,
        "halt_fire_rate": round(halt_fires / total, 3) if total else 0.0,
        "current_halt_reason": state.loop.halt_reason,
        "mid_loop_interventions": len(interventions),
        "chain_consume_failures": len(consume_fails),
        "release_owner": state.release_owner,
    }


# --- Governance health ----------------------------------------------------


def _parse_dt(s: Any) -> Optional[datetime]:
    """Best-effort parse of an ISO-ish string into a UTC datetime."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=timezone.utc)
    s = str(s).strip()
    if not s:
        return None
    # Strip a trailing Z
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        # Try common date-only and date+time forms
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%Y %H:%M"):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


def _governance_health(
    state: ls.State, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    sent_rows = [r for r in rows if str(r.get("Status", "")).strip() in SENT_STATUSES]
    staged_rows = [r for r in rows if str(r.get("Status", "")).strip() in STAGED_STATUSES]
    killed_rows = [r for r in rows if str(r.get("Status", "")).strip() in KILLED_STATUSES]
    queued_rows = [r for r in rows if str(r.get("Status", "")).strip() in QUEUED_STATUSES]

    total_decided = len(sent_rows) + len(killed_rows)

    # sent_as_staged_ratio: of everything the human has DECIDED on (sent or
    # killed), how much went out as drafted vs went to the trash. High ratio
    # = rubber-stamping; low ratio = active gatekeeping. The paper does NOT
    # prescribe a threshold; we just surface the number.
    if total_decided > 0:
        sent_ratio = round(len(sent_rows) / total_decided, 3)
        killed_ratio = round(len(killed_rows) / total_decided, 3)
    else:
        sent_ratio = 0.0
        killed_ratio = 0.0

    # mean_time_to_first_send: from the loop's last calendar block to the
    # earliest Posted At timestamp in the tracker. This measures how long
    # the human took to actually sit down at the calendar block. A very
    # short time + a very high sent_ratio is the worst combination —
    # bot reflexes from a human-stamped queue.
    cal_dt = _parse_dt(state.handoff.last_calendar_block_utc)
    response_secs: list[float] = []
    if cal_dt:
        for r in sent_rows:
            posted_at = _parse_dt(
                r.get("Posted At") or r.get("Posted at") or r.get("First Touch")
            )
            if posted_at and posted_at >= cal_dt:
                response_secs.append((posted_at - cal_dt).total_seconds())

    mean_response_min = (
        round(statistics.mean(response_secs) / 60.0, 1) if response_secs else None
    )

    return {
        "batches_sent": len(sent_rows),
        "batches_staged": len(staged_rows),
        "batches_killed": len(killed_rows),
        "batches_queued": len(queued_rows),
        "sent_as_staged_ratio": sent_ratio,
        "killed_ratio": killed_ratio,
        "decided_total": total_decided,
        "mean_response_minutes_to_first_send": mean_response_min,
        "response_observations": len(response_secs),
        "last_calendar_block_utc": state.handoff.last_calendar_block_utc,
        "last_calendar_block_id": state.handoff.last_calendar_block_id,
    }


# --- Top-level report ------------------------------------------------------


def build_report(loopdir: Path, tracker_path: Path) -> dict[str, Any]:
    state = ls.load(loopdir)
    ticks = _read_tick_logs(loopdir)
    audit = _read_audit_log(loopdir)
    rows = _read_tracker_rows(tracker_path)

    return {
        "project": state.project,
        "generated_utc": ls.utcnow_iso(),
        "loop_state_dir": str(loopdir),
        "tracker_path": str(tracker_path),
        "engineering_health": _engineering_health(state, ticks, audit),
        "governance_health": _governance_health(state, rows),
    }


def _fmt_pct(n: float) -> str:
    return f"{n * 100:5.1f}%"


def render_text(report: dict[str, Any]) -> str:
    eng = report["engineering_health"]
    gov = report["governance_health"]
    lines = [
        f"Synergy Loop Audit — {report['project']}",
        f"  generated:     {report['generated_utc']}",
        f"  loopdir:       {report['loop_state_dir']}",
        f"  tracker:       {report['tracker_path']}",
        "",
        "Engineering health (was the loop machinery exercised?)",
        f"  release_owner:           {eng['release_owner']}",
        f"  ticks total:             {eng['ticks_total']}",
        f"  verifier hard-stops:     {eng['verifier_hard_stops']:4d}   ({_fmt_pct(eng['verifier_hard_stop_rate'])} of ticks)",
        f"  transient retries:       {eng['transient_retries']:4d}",
        f"  goal-met ticks:          {eng['goal_met_count']:4d}",
        f"  halt fires (any cause):  {eng['halt_fires']:4d}   ({_fmt_pct(eng['halt_fire_rate'])} of ticks)",
        f"  current halt_reason:     {eng['current_halt_reason'] or '(none)'}",
        f"  mid-loop interventions:  {eng['mid_loop_interventions']:4d}",
        f"  chain consume failures:  {eng['chain_consume_failures']:4d}   (suspicious if non-zero)",
        "",
        "Governance health (was the human actually overriding?)",
        f"  batches sent:            {gov['batches_sent']:4d}",
        f"  batches staged:          {gov['batches_staged']:4d}",
        f"  batches killed:          {gov['batches_killed']:4d}",
        f"  batches queued:          {gov['batches_queued']:4d}",
        f"  decided (sent+killed):   {gov['decided_total']:4d}",
        f"  sent-as-staged ratio:    {_fmt_pct(gov['sent_as_staged_ratio'])}   (high = rubber-stamp risk)",
        f"  kill ratio:              {_fmt_pct(gov['killed_ratio'])}",
        f"  last calendar block:     {gov['last_calendar_block_utc'] or '(none)'}",
        f"  mean response to first send:  "
        + (f"{gov['mean_response_minutes_to_first_send']} min"
           f" (n={gov['response_observations']})"
           if gov["mean_response_minutes_to_first_send"] is not None
           else "(no calendar block + sent row yet)"),
        "",
        "Per the Governed Autonomy paper, these two halves are read together:",
        "  - engineering metrics alone can't tell you the human is paying attention",
        "  - governance metrics alone can't tell you the system is even working",
        "  - both must be inside their own thresholds, and the thresholds are yours to set",
    ]
    return "\n".join(lines)


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("loopdir", type=Path)
    p.add_argument("tracker_path", type=Path)
    p.add_argument(
        "--json",
        action="store_true",
        help="Emit JSON instead of the formatted text report.",
    )
    args = p.parse_args(argv)

    try:
        report = build_report(args.loopdir, args.tracker_path)
    except ls.StateNotFoundError as e:
        print(f"ERROR: no state.json at {e}", file=sys.stderr)
        return 30
    except ls.StateSchemaError as e:
        print(f"ERROR: state.json schema invalid: {e}", file=sys.stderr)
        return 30

    if args.json:
        print(json.dumps(report, indent=2, sort_keys=True))
    else:
        print(render_text(report))
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
